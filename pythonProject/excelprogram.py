import openpyxl
filename="file:///C:/Users/maria/OneDrive/Documents/tkintersignin.xlsx"
if (filename):
    wb=openpyxl.workbook(filename)
    if 'Data' in wb.get_sheet_names():
      pass
    else:
        wb.create_sheet(index=0, title='Data')
else:
   wb = openpyxl.Workbook()
   wb.create_sheet(index=0, title='Data')
   wb.save(filename)

wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Data')

def add_username(name,password):
   ws = wb.active
   first_column = ws['A']
   second_column = ws['B']
   col_len1 = str(len(first_column)+1)
   col_len2 = str(len(second_column)+1)
   sheet['A' + col_len1] = name
   sheet['B' + col_len2] = password
   wb.save(filename)

if (sheet['A1'].value == 'Username') and (sheet['B1'].value == 'Password'):
   pass
else:
   sheet['A1'] = 'Username'
   sheet['B1'] = 'Password'